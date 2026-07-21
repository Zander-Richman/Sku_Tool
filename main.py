import os
import tempfile
import cloudmersive_barcode_api_client
from cloudmersive_barcode_api_client.rest import ApiException
import shutil
import pytesseract
import time
from PIL import Image
from pillow_heif import register_heif_opener
import paramiko
from win32com.client import Dispatch
import sys
from google.cloud import vision
import re



# -------SETTINGS-------

if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))  # Get the directory of the current script
EXCEL_PATH = os.path.join(BASE_DIR, 'BORG_Shared', 'inventory_images_record.xlsx')  # Path to save the Excel log file
SOURCE_FOLDER = os.path.join(BASE_DIR, "product_input")  # Default source folder
OUTPUT_FOLDER = os.path.join(BASE_DIR, "product_output")  # Default output folder
SUPPLIER_PREFIXES = {"KAWASAKI": "KA",
                     "YAMAHA": "YA",
                     "SUZUKI": "SU",
                     "KTM": "KT",
                     "POLARIS": "PO",
                     "BRP": "SD",
                     "CF MOTO": "CF",
                     "ARCTIC CAT": "AC",
                     "HONDA": "HO"}
BARCODE_SUPPLIERS = ["KAWASAKI", "YAMAHA", "SUZUZI", "HONDA", "POLARIS", "KTM", "BRP", "CF MOTO"]
TEXT_SUPPLIERS = ["ARCTIC CAT"]


# Path to the logo image.... may need this later if we decide to add it
LOGO_PATH = os.path.join(BASE_DIR, "worldofpowersports_logo.png")  

# Google vision client creation
vision_client = vision.ImageAnnotatorClient(
                client_options = {"api_key": os.getenv("GOOGLE_VISION_KEY")}
            )
# Cloudmersive barcode API client creation
configuration = cloudmersive_barcode_api_client.Configuration()
configuration.api_key['Apikey'] = os.getenv("CLOUDMERSIVE_KEY")
api_instance = cloudmersive_barcode_api_client.BarcodeScanApi(
    cloudmersive_barcode_api_client.ApiClient(configuration)
)


def create_shortcut():
    try:
        desktop = os.path.join(os.environ['USERPROFILE'], 'OneDrive', 'Desktop')
        shortcut_path = os.path.join(desktop, 'BORG.lnk')
        print("finding shortcut...")

        if not os.path.exists(shortcut_path):
            shell = Dispatch('WScript.Shell')
            shortcut = shell.CreateShortCut(shortcut_path)
            shortcut.Targetpath = BASE_DIR
            shortcut.IconLocation = os.path.join(BASE_DIR, 'borg_icon.ico')
            shortcut.save()
            print("BORG shortcut created on Desktop! :)")
        else:
            print("BORG shortcut already exists on Desktop! :)")
    except Exception as e:
        print(f"Could not create shortcut: {e}")


def setup_ftp_credentials():
        print("FTP credentials not found -- first time setup required.")
        print("Please ask your supervisor for the FTP credentials.")
        print("")
        host = input("Enter FTP host: ").strip()
        user = input("Enter your FTP username: ").strip()
        password = input("Enter your FTP password: ")
        path = input("Enter the FTP path: ")

        os.system(f'setx FTP_HOST "{host}"')
        os.system(f'setx FTP_USER "{user}"')
        os.system(f'setx FTP_PASSWORD "{password}"')
        os.system(f'setx FTP_PATH "{path}"')

        print("FTP credentials saved! Please restart application for changes to take effect.")
        input("Press enter to exit...")
        raise SystemExit(0)


def first_time_setup():
    os.makedirs(SOURCE_FOLDER, exist_ok=True)
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    create_shortcut()

    if not os.path.exists(EXCEL_PATH):
        print("ERROR: the file 'inventory_images_record.xlsx' does not exist in the BORG folder." \
        "You are not allowed to run this tool until you add this file.")
        input("Please seek assistance from your supervisor to obtain the 'inventory_images_record.xlsx' file and place it in the BORG folder, or await further instructions." \

            "Press enter to exit the application once you have read this message. ")
        raise SystemExit(0)
    
    if not os.getenv("FTP_HOST"):
        setup_ftp_credentials()

    print("product_input folder is ready")
    print("product_output folder is ready")
    print("Excel log file is ready")
    print("FTP credentials have been entered")
    setup_confirmation = input("If you have not done so yet, please insert pictures into the input folder. " \
        "If pictures from the previous session are still in the folder, please type 'clear' to clear the input folder. " \
        "If your folder is cleared but there are no new photos, type 'exit' to exit the program and insert your pictures into the folder. " \

        "Otherwise, press Enter to continue. ")
    if setup_confirmation == "":
        return
    elif setup_confirmation.lower().strip() == "exit":
        raise SystemExit(0)
    elif setup_confirmation.lower().strip() == "clear":
        confirm_clear = input(f"Are you sure you want to clear the source folder ({SOURCE_FOLDER})? This will delete all original photos. Type 'yes' to confirm: ")
        if confirm_clear.lower().strip() == 'yes':
            clear_input_folder()
            input("Input folder cleared! Please add new photos and restart BORG.")
            raise SystemExit(0)


def get_jpgs(folder):
    valid_extentions = ['.jpg', '.jpeg']  # Add more valid extensions if needed
    jpgs= []
    for file in os.listdir(folder):
        ext = os.path.splitext(file)[1].lower()
        if ext in valid_extentions:
            jpgs.append(os.path.join(folder, file))
    jpgs.sort()
    return jpgs


def get_all_images(folder):
    for file in os.listdir(folder):
        print(f"file: {file}")
    valid_extentions = ['.jpg', '.jpeg', '.png', '.heic', '.webp', '.aae']  # Add more valid extensions if needed
    images = []
    for file in os.listdir(folder):
        ext = os.path.splitext(file)[1].lower()
        if ext in valid_extentions:
            images.append(os.path.join(folder, file))
    images.sort()
    return images
    

def convert_to_jpg(image_path):
    if image_path.lower().endswith('.jpg') or image_path.lower().endswith('.jpeg'):
        return image_path  # No conversion needed
    
    jpg_path = os.path.splitext(image_path)[0] + '.jpg'
    img = Image.open(image_path)
    img.convert('RGB').save(jpg_path, 'JPEG')
    print(f"Converted {image_path} to {jpg_path}")
    return jpg_path


def detect_barcode(image_path):
    try: 
        time.sleep(1)  # Add a delay to avoid hitting rate limits
        with Image.open((image_path)) as img:
            img = img.convert('RGB')
            temp = tempfile.NamedTemporaryFile(suffix='.jpg', delete=False)
            temp_path = temp.name 
            temp.close()
            img.save(temp_path, 'JPEG')
        response = api_instance.barcode_scan_image(temp_path)
        os.remove(temp.name)

        if response.successful:
            return response.raw_text
        return None
    except ApiException as e:
        return None
    except Exception as e:
        input(f"Connection timed out during barcode detection: {e}")
        return None


def detect_text(image_path, supplier, retries=2):
    google_vision_failed = False
    for attempt in range(retries): 
        try:    
            with open(image_path, 'rb') as f:
                content = f.read()
            image = vision.Image(content=content)
            response = vision_client.text_detection(image=image)


            if response.error.message:
                print (f"Vision API error: {response.error.message}")
                return None
            
            if response.text_annotations:
                text = response.text_annotations[0].description.upper()
                result = extract_sku_from_text(text, supplier)
                if result:
                    return result
        
        except Exception as e:
            print (f"Google Vision attempt {attempt + 1} failed: {e}")
            google_vision_failed = True
            time.sleep(2)
            
    if google_vision_failed == True:
        print("Falling back to Tesseract OCR...")
        try:
            img = Image.open(image_path).convert('L')
            text = pytesseract.image_to_string(img, config=r'--psm 6').upper()
            return extract_sku_from_text(text, supplier)
        except Exception as e:
            print(f"Tesseract also failed: {e}")
            return None


def extract_sku_from_text(text, supplier):
    if supplier == "Arctic Cat":
        match = re.search(r'\b[A-Z0-9]{4}-[A-Z0-9]{3}\b', text)
        if match:
            return match.group()
            
    else:
        lines = text.splitlines()
        lines = [line.strip() for line in lines if line.strip()]
        if any('SKU' in line for line in lines):
            sku_index = next(i for i, line in enumerate(lines) if 'SKU' in line)
            if sku_index + 1 < len(lines):
                sku = lines[sku_index + 1].replace(' ', '').strip()
                sku = ''.join(c for c in sku if c.isalnum() or c in '-_')
                return sku
        return None


def parse_sku(raw, supplier):
    if supplier == "KAWASAKI":
        raw = raw.replace('-','')
        if len(raw) == 9:
            return f"{raw[0:5]}-{raw[5:9]}"
        elif len(raw) == 8:
            return f"{raw[0:5]}-{raw[5:8]}"
        elif len(raw) == 12:
            return f"{raw[0:5]}-{raw[5:9]}-{raw[9:12]}"
    
    elif supplier == "YAMAHA":
        raw =  raw.replace('-', '')
        if len(raw) == 12:    
            if raw.startswith('9'):
                return f"{raw[0:5]}-{raw[5:10]}-{raw[10:12]}"
            else:
                return f"{raw[0:3]}-{raw[3:8]}-{raw[8:10]}-{raw[10:12]}"
    elif supplier == "SUZUKI":
        parts = raw.split('-')
        if all(c == '0' for c in parts[-1]):
            return '-'.join(parts[:-1])
        return raw
    
    elif supplier == "CF MOTO":
        if raw.endswith('*1') or raw.endswith('*2'):
            return raw[:-2].strip('-')
        return raw
    
    elif supplier == "BRP":
        digits = raw[:9]
        #BRP barcodes are 9 digits long, so if the raw text is 9 digits, we can return it as the SKU. Otherwise, we return None to indicate that it doesn't match the expected format.
        if len(raw) == 9 and digits.isdigit():
            return digits
        return None # doesn't match BRP format, not a real barcode
    
    else:
        return raw


def is_divider(image_path, supplier):
    sku = None

    if supplier in BARCODE_SUPPLIERS:
        raw = detect_barcode(image_path)
        if raw:
            sku = parse_sku(raw, supplier)
        
    elif supplier in TEXT_SUPPLIERS:
        raw = detect_text(image_path, supplier)
        if raw:
            sku = raw
    
    elif supplier == "Other":
        raw = detect_text(image_path, supplier)
        print(f"Text detection result for {os.path.basename(image_path)}: {raw}")
        if raw:
            sku = raw

    if sku: 
        if supplier != "Other":
            prefix = SUPPLIER_PREFIXES.get(supplier, "")
            sku = prefix + sku
        
        print(f"Detected SKU: {sku} in {os.path.basename(image_path)}")


        return sku
    
    return None


def process_photos(photos, supplier):
    current_sku= None
    groups = {}
    flagged = {}
    failed = []

    for photo in photos: 
        sku = is_divider(photo, supplier) 
        if sku:
            current_sku = sku
            groups[sku] = []
            print(f"New groups started: {sku}")
        elif current_sku:
            groups[current_sku].append(photo)
        else: 
            failed.append(os.path.basename(photo))

    for sku in list(groups.keys()):
        if len(groups[sku]) > 6:
            flagged[sku] = groups.pop(sku)
            print(f"Flagged {sku} pulled from upload - {len(flagged[sku])} photos. It is possible that a divider photo was missed. Needs manual review.")

    
    
    if failed:
        print(f"\n {len(failed)} photos could not be grouped: ")
        for f in failed:
            print(f" - {f}")
        print("Please manually assign these photos to the correct SKU folder.")

    return groups, flagged, failed


def export_folders(groups):
    # Create output folder if it doesn't exist
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)

    for sku, photos in groups.items():
        sku_folder = os.path.join(OUTPUT_FOLDER, sku)
        os.makedirs(sku_folder, exist_ok=True)

        for i, photo in enumerate(photos, start=1):
            new_filename = f"{sku}__{i}.jpg"
            destination = os.path.join(sku_folder, new_filename)
            shutil.copy(photo, destination)

        if os.path.exists(LOGO_PATH):
            new_filename = f"{sku}__{len(photos) + 1}.jpg"
            logo_destination = os.path.join(sku_folder, new_filename)
            shutil.copy(LOGO_PATH, logo_destination)

    print(f"\nDONE! {len(groups)} SKU folders created in {OUTPUT_FOLDER}")


def clear_input_folder():
    for file in os.listdir(SOURCE_FOLDER):
        file_path = os.path.join(SOURCE_FOLDER, file)
        if os.path.isfile(file_path):
            os.remove(file_path)
    print(f"Cleared all files in {SOURCE_FOLDER}")


def upload_to_ftp(groups):
    FTP_HOST = os.getenv("FTP_HOST")
    FTP_USER = os.getenv("FTP_USER")
    FTP_PASS = os.getenv("FTP_PASSWORD")
    FTP_PATH = os.getenv("FTP_PATH")

    print(f"Uploading to FTP server: {FTP_HOST}")

    try:
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(hostname=FTP_HOST, port=22, username=FTP_USER, password=FTP_PASS)
        print(f"Connected to SSH server: {FTP_HOST}")
        sftp = client.open_sftp()
        sftp.chdir(FTP_PATH)
        print(f"Changed directory to: {FTP_PATH}")

        for sku in groups.keys():
            local_folder = os.path.join(OUTPUT_FOLDER, sku)
            if os.path.isdir(local_folder):
                for file in os.listdir(local_folder):
                    if file.endswith('.jpg'):
                        local_path = os.path.join(local_folder, file)
                        remote_path = f"{FTP_PATH}/{file}"
                        sftp.put(local_path, remote_path)
                        print(f"Uploaded: {file}")
        sftp.close()
        client.close()
        print("All files uploaded successfully!")
        return True
    except Exception as e:
        print(f"Failed to upload to FTP server: {e}")
        return False


def update_excel_log(groups, imported=False):
    import openpyxl

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    except PermissionError:
        print("Please close out of excel and restart the program. Once restarted, type 'update' when prompted to do so. :)")
        return
    
    for row in ws.iter_rows(min_row=2):
        sku_cell = row[0]
        photos_taken_cell = row[4]
        photos_imported_cell = row[5]

        if sku_cell.value in groups:
            photos_taken_cell.value = "x"
            if imported == True:
                photos_imported_cell.value = "x"
                print(f"Marked SKU {sku_cell.value} as photographed and imported")
            else:
                print(f"Marked SKU {sku_cell.value} as photographed but not imported")


    wb.save(EXCEL_PATH)
    print(f"Excel log updated at: {EXCEL_PATH}")


def add_to_excel_log(sku, imported=False):
    import openpyxl

    try:
        wb = openpyxl.load_workbook(EXCEL_PATH)
        ws = wb.active
    except PermissionError:
        print("Please close out of excel and restart the program. Once restarted, type 'update' when prompted to do so. :)")
        return

    # Check if SKU already exists in the log
    for row in ws.iter_rows(min_row=2):
        sku_cell = row[0]
        if sku_cell.value == sku:
            print(f"SKU {sku} already exists in the Excel log.")
            return

    # If SKU does not exist, add it to the next available row
    last_row = 2
    for row in ws.iter_rows(min_row = 1, max_col = 1):
        if row[0].value:
            last_row = row[0].row

    next_row = last_row + 1
        
    ws.cell(row=next_row, column=1, value=sku)  # SKU
    ws.cell(row=next_row, column=5, value="x")  # Photos Taken
    ws.cell(row=next_row, column=2, value="")  # Inventory count (placeholder)
    ws.cell(row=next_row, column=3, value="") # Bin Location (placeholder)
    if imported:
        ws.cell(row=next_row, column=6, value="x")  # Photos Imported

    wb.save(EXCEL_PATH)
    print(f"Added SKU {sku} to Excel log at row {next_row}.")

#step 1: set up folders and check for Excel log file then start the process
register_heif_opener()
first_time_setup()

# Step 2: Ask for supplier
print("Suppliers: Kawasaki, Yamaha, Suzuki, Honda, Polaris, KTM, BRP, CF Moto, Arctic Cat, Other")
supplier = input("Enter your supplier brand for this session: ").strip().upper()


if supplier == "Other":
    print("Note: 'Other' is intended for occasional use with mixed or unlisted brands.")
    print("Please write the full SKU including supplier code on your divider card.")
    print("This program will read it exactly as written.")

else:
    if supplier not in SUPPLIER_PREFIXES:
        input("Supplier not recognized! Please choose from the list.")
        raise SystemExit(0)
    print(f"Processing {supplier} products - {SUPPLIER_PREFIXES[supplier]} prefix will be applied.")


start = input("Press Enter to start processing photos, type 'update' to update the Excel Log")


# Step 2 — Convert all photos to JPG
if start == "":
    print("BORG is running!!!")
    print(f"Looking for photos in: {SOURCE_FOLDER}")
    
    print("Converting all photos to JPG...")
    for photo in get_all_images(SOURCE_FOLDER):
        convert_to_jpg(photo)

    # Step 3 — Get only JPGs
    photos = get_jpgs(SOURCE_FOLDER)
    print(f"Found {len(photos)} JPG images")

    # Step 4 — Process them
    groups, flagged, failed = process_photos(photos, supplier)
    print(f"\nFound {len(groups)} SKUs")
    for sku, photos in groups.items():
        print(f"{sku}: {len(photos)} photos")

    # Step 5 — create the folders with the images in the output folder
    export_folders(groups)

    # Step 6 - Upload to FTP
    upload_success = upload_to_ftp(groups)


    # Step 7 - address flagged SKUs
    if flagged:
        for sku, photos_list in flagged.items():
            print(f"SKU: {sku} has {len(photos_list)} photos. Please refer to the input folder and determine if a divider photo was missed.")
            decision =input("Should this SKU group be uploaded as-is (yes/no)?")

            if decision.lower().strip() == "yes":
                groups[sku] = photos_list
                export_folders({sku: photos_list})
                upload_to_ftp({sku: photos_list})
                add_to_excel_log(sku, imported=True)
            else:
                review_folder = os.path.join(OUTPUT_FOLDER, "NEEDS_REVIEW", sku)
                os.makedirs(review_folder, exist_ok=True)
                for i, photo in enumerate(photos_list, start=1):
                    shutil.copy(photo, os.path.join(review_folder, f"{sku}__{i}.jpg"))


    # Step 7 - Update Excel log
    for sku in groups.keys():
        add_to_excel_log(sku, imported=upload_success)
    update_excel_log(groups, imported=upload_success)
    
    
    # Step 8 — Clear source folder
    confirm_clear = input(f"Do you want to clear the source folder ({SOURCE_FOLDER})? This will delete all original photos. Type 'yes' to confirm: ")
    if confirm_clear.lower() == 'yes':
        clear_input_folder()


elif start.lower() == "update":
    groups = {folder: [] for folder in os.listdir(OUTPUT_FOLDER) 
              if os.path.isdir(os.path.join(OUTPUT_FOLDER, folder))}
    did_upload = input("Did you already upload the photos to the FTP server? If yes, type 'yes' and press Enter to update the Excel log with imported status. If not, just press Enter to update without imported status: ")
    update_excel_log(groups, imported = (did_upload.lower() == 'yes'))
